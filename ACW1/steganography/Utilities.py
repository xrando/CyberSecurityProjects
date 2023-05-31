#pip install python-docx openpyxl
import docx
import openpyxl

def read_file_content(file_path):
    file_extension = file_path.split('.')[-1].lower()

    if file_extension == 'txt':
        with open(file_path, 'r') as file:
            content = file.read()
        return content
    elif file_extension == 'docx':
        doc = docx.Document(file_path)
        content = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
        return content
    elif file_extension == 'xlsx':
        workbook = openpyxl.load_workbook(file_path)
        content = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                content.append('\t'.join(str(cell) for cell in row))
        return '\n'.join(content)
    else:
        return None

