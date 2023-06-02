from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np

class ExcelSteganography:
    # initialize class instance for storing file path, loading workbook & setting active sheet
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = load_workbook(file_path)
        self.sheet = self.workbook.active

    # converting the data to binary format as string
    # checks the type of the data and perform the conversion accordingly
    def toBinary(self, data):
        if isinstance(data, str):
            return ''.join([format(ord(i), "08b") for i in data])
        elif isinstance(data, bytes):
            return ''.join([format(i, "08b") for i in data])
        elif isinstance(data, np.ndarray) or isinstance(data, list) or isinstance(data, tuple):
            return [format(i, "08b") for i in data]
        elif isinstance(data, int) or isinstance(data, np.uint8):
            return format(data, "08b")
        else:
            raise TypeError("Type is not supported.")
    
    # encoding function
    def encode(self, message, bit):
        # Adds "#####" to the message to mark the end of the message
        message += "#####"  
        # Converts the message to binary using the 'toBinary' method
        binaryMessage = self.toBinary(message)  
        # Initializes the index variable to 0
        index = 0  
        # Calculates the length of the binary message
        length = len(binaryMessage)  
        # Calculates the maximum allowed length based on the sheet size and bit value
        maxLength = (self.sheet.max_row - 1) * self.sheet.max_column * bit  
        # Checks if the length of the message exceeds the maximum allowed length
        if length > maxLength:  
            raise ValueError(f"[!] The message with length of ({length}) has exceeded the max length of {maxLength}, please change the message length or bit to change.")
        # Iterates over each row in the active sheet
        for row in self.sheet.iter_rows():  
            # If the index position exceeds the length of the message, breaks the loop
            if index >= length:  
                break
            # For each cell in the row
            for cell in row:  
                # Checks if the index is within the message length
                if index < length:  
                    # Get current color and fill type of the cell
                    try:
                        fill = cell.fill
                        color = fill.fgColor.rgb
                        patternType = fill.patternType
                    except AttributeError:
                        # Default fill background color of white if not specified
                        color = "FFFFFFFF"
                        patternType = "none"
                    # Converts the color to binary and pads with zeros to 24 bits (RGB color)
                    fillBinary = bin(int(color, 16))[2:].zfill(24)  
                    data = ""
                    # Iterates 'bit' number of times
                    for i in range(bit):  
                        # Checks if the index is within the message length
                        if index < length:  
                            # Appends the binary value from the message to the 'data' variable
                            data += binaryMessage[index]  
                            # Increments the index
                            index += 1  
                        else:
                            # If the index exceeds the message length, appends the remaining bits from the color
                            data += fillBinary[-bit + i]  
                    # Changes the 'fillBinary' by replacing the last 'bit' number of bits with 'data'
                    fillBinary = fillBinary[:-bit] + data  
                    # Extracts the red component from the binary fill value
                    red = int(fillBinary[:8], 2)  
                    # Extracts the green component from the binary fill value
                    green = int(fillBinary[8:16], 2)  
                    # Extracts the blue component from the binary fill value
                    blue = int(fillBinary[16:], 2)  
                    # Creates a new fill with the updated RGB color
                    fill = PatternFill(patternType=patternType, fgColor=f"FF{red:02X}{green:02X}{blue:02X}")  
                    # Updates the cell background color with the new fill
                    cell.fill = fill  
                else:
                    # If the index exceeds the message length, breaks the loop
                    break  
                
        # Returns the modified workbook
        return self.workbook  

    # encoding function with the payload
    def encode_with_payload(self, payload_file_path, bit):
        # Opens the payload file in read mode
        with open(payload_file_path, 'r') as file:  
            # Reads the contents of the file and assigns it to the 'message' variable
            message = file.read()  

        # Calls the 'encode' method with the 'message' and 'bit' as arguments and returns the result
        return self.encode(message, bit)  

    # decoding function
    def decode(self, bit):
        # To store the binary representation of the hidden message
        binaryMessage = ""  
        # To store the decoded message
        decodedMessage = ""  
        # Iterates over each row in the active sheet
        for row in self.sheet.iter_rows():  
            # Iterates over each cell in the row
            for cell in row:  
                # Retrieves the fill object of the cell
                fill = cell.fill  
                # Retrieves the RGB color value of the cell's fill
                color = fill.fgColor.rgb  
                # Extracts the red component from the RGB color value
                red = int(color[2:4], 16)  
                # Extracts the green component from the RGB color value
                green = int(color[4:6], 16)  
                # Extracts the blue component from the RGB color value
                blue = int(color[6:], 16)  
                # Converts the RGB color components to an 8-bit binary string
                binary_color = f"{red:08b}{green:08b}{blue:08b}"  
                # Appends the least significant 'bit' number of bits from the binary color string to the 'binaryMessage'
                binaryMessage += binary_color[-bit:]  
                # Checks if there are enough bits in 'binaryMessage' to form a complete character (8 bits)
                if len(binaryMessage) >= 8:  
                    # Converts the first 8 bits of 'binaryMessage' to a character and appends it to the 'message'
                    decodedMessage += chr(int(binaryMessage[:8], 2))  
                    # Checks if the 'message' ends with the marker "#####"
                    if decodedMessage.endswith("#####"):  
                        # If the marker is found, returns the 'message' without the marker
                        return decodedMessage[:-5]  
                    # Removes the first 8 bits from 'binaryMessage' to process the next character
                    binaryMessage = binaryMessage[8:]  

    # Saving the modified workbook to a specified file path. 
    def save(self, save_file_path):
        # Utilizes the save method of the openpyxl workbook object to perform the actual saving operation.
        self.workbook.save(save_file_path)

# initialize obj with the cover excel file
obj = ExcelSteganography('files/cover.xlsx')

# encode with the payload
encoded_workbook = obj.encode_with_payload('files/payload.txt', 3)

# save the encoded file
obj.save('files/encoded.xlsx')

# decode the encoded file
decoded_msg = obj.decode(3)
print(decoded_msg)